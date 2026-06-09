#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
统计参数工作簿 / Statistical-parameters workbook
科学问题/目标: 把所有 hazard 模型的系数、HR、CI、p、AIC、Akaike weight、
  诊断与预测统计参数汇总为一个可编辑的 Excel(多 sheet)，便于复核与复用。
Aggregate every persisted model parameter into one editable .xlsx.
输入: results/tables/*.csv, results/diagnostics/*.csv, results/forecasts/*.csv
输出: output/doc/bird_hazard_all_statistics.xlsx
运行: python3 code/55_stats_workbook.py
"""
from pathlib import Path
import pandas as pd
import numpy as np

ROOT = Path(".").resolve()
T = ROOT / "results" / "tables"
D = ROOT / "results" / "diagnostics"
F = ROOT / "results" / "forecasts"
OUT = ROOT / "output" / "doc" / "bird_hazard_all_statistics.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

def rd(p):
    try:
        return pd.read_csv(p)
    except Exception as e:  # noqa
        print(f"  [skip] {p.name}: {e}")
        return None

# sheet_name -> (path, description)
SHEETS = [
    ("README", None, None),
    ("province_v2_coefs", T/"table_province_v2_coefs.csv",
     "Conservative (v2) province M0-M4 coefficients, 4 effort specs; HR/CI/p"),
    ("province_v2_AIC", T/"table_province_v2_aic.csv",
     "v2 province AIC ladder M0-M4 x 4 specs"),
    ("AIC_akaike_weights", T/"table_aic_akaike_weights.csv",
     "Akaike weights per spec (M0-M4)"),
    ("M5_offset_test", T/"table_m5_offset_summary.csv",
     "M4(interaction) vs M5(offset): moderator vs scaling, v2+v3"),
    ("endogeneity_lag", T/"table_effort_lag_refit.csv",
     "Effort lagged 1yr (province-year) interaction; endogeneity defense"),
    ("multiscale_pref_coefs", T/"table_prefecture_coefs.csv",
     "Prefecture refit M0-M4 fixed effects"),
    ("multiscale_county_coefs", T/"table_county_coefs.csv",
     "County refit M0-M4 fixed effects"),
    ("multiscale_AIC", T/"table_prefecture_county_aic.csv",
     "Prefecture/county AIC + Akaike weight"),
    ("v3_allspecs_coefs", T/"table_province_v3_all_specs_coefs.csv",
     "Relaxed (v3) province M0-M4 coefficients, 4 specs"),
    ("v3_allspecs_AIC", T/"table_province_v3_all_specs_aic.csv",
     "v3 province AIC + Akaike weights"),
    ("v3_three_scale", T/"table_v3_three_scale_summary.csv",
     "v3 province/prefecture/county interaction HR/CI/p"),
    ("reconciliation_v1v2v3", T/"table_province_v1_v2_v3_reconciliation.csv",
     "v1/v2/v3 headline interaction reconciliation"),
    ("spatial_block_CV", T/"table_spatial_block_cv.csv",
     "Spatial-block(250km) vs random 5-fold AUC"),
    ("morans_I_residuals", D/"table_morans_i_residuals.csv",
     "M4 residual Moran's I at 100/250/500 km"),
    ("RF_importance_v2", T/"table_rf_importance_v2.csv",
     "Random-forest permutation importance (v2)"),
    ("RF_v2_v3_rank_shift", T/"table_v2_v3_rf_comparison.csv",
     "RF importance rank shift v2 -> v3"),
    ("xgb_cv_v3", T/"table_xgb_cv_v3.csv",
     "XGBoost 5-fold CV-AUC (v3 relaxed)"),
    ("riskset_v3_attrition", D/"table_riskset_v3_attrition.csv",
     "Risk-set attrition funnel raw -> v3"),
    ("future_province_glmmTMB", F/"table_province_future_glmmTMB.csv",
     "Province future hazard (glmmTMB) SSP x year"),
    ("future_province_xgboost", F/"table_province_future_xgboost.csv",
     "Province future hazard (XGBoost) SSP x year"),
]

# Build a tidy "headline interaction" summary sheet from the three key tables
def headline_summary():
    rows = []
    v2 = rd(T/"table_province_v2_coefs.csv")
    if v2 is not None:
        s = v2[(v2.model == "M4") & (v2.term == "climate_z:effort_z")]
        for _, r in s.iterrows():
            rows.append(["v2 conservative (concurrent effort)", r.spec_label,
                         r.hr, r["hr.low"], r["hr.high"], r["p.value"]])
    lag = rd(T/"table_effort_lag_refit.csv")
    if lag is not None:
        for _, r in lag.iterrows():
            rows.append(["v2 lagged effort (t-1)", r.spec_label,
                         r.hr, r["hr.low"], r["hr.high"], r["p.value"]])
    v3 = rd(T/"table_province_v3_all_specs_coefs.csv")
    if v3 is not None:
        s = v3[(v3.model == "M4") & (v3.term == "climate_z:effort_z")]
        for _, r in s.iterrows():
            rows.append(["v3 relaxed (concurrent effort)", r.spec_label,
                         r.hr, r["hr.low"], r["hr.high"], r["p.value"]])
    df = pd.DataFrame(rows, columns=["dataset", "effort_spec", "HR",
                                     "CI_low", "CI_high", "p_value"])
    return df

readme = pd.DataFrame(
    [[n, (d or "")] for (n, _, d) in SHEETS if n != "README"],
    columns=["sheet", "description"])

with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    readme.to_excel(xw, sheet_name="README", index=False)
    headline_summary().to_excel(xw, sheet_name="headline_interaction", index=False)
    for name, path, _ in SHEETS:
        if name == "README" or path is None:
            continue
        df = rd(path)
        if df is not None:
            df.to_excel(xw, sheet_name=name[:31], index=False)

# light column-width autofit
from openpyxl import load_workbook
wb = load_workbook(OUT)
for ws in wb.worksheets:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 48)
wb.save(OUT)
print(f"[55] wrote {OUT}")
print(f"[55] sheets: {len(wb.sheetnames)} -> {wb.sheetnames}")
